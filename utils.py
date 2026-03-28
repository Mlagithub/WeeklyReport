from datetime import datetime, timedelta
from io import BytesIO

from bs4 import BeautifulSoup
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class DateRange:
    TIME_RANGES = {
        "this_week": "本周",
        "last_week": "上周",
        "this_month": "本月",
        "this_quarter": "本季度",
        "this_year": "本年",
    }

    @staticmethod
    def get_today():
        return datetime.today().date()

    @staticmethod
    def this_week():
        today = DateRange.get_today()
        start_date = today - timedelta(days=today.weekday())
        end_date = today
        return start_date, end_date

    @staticmethod
    def last_week():
        today = DateRange.get_today()
        start_date = today - timedelta(days=today.weekday() + 7)
        end_date = start_date + timedelta(days=6)  # 上周日
        return start_date, end_date

    @staticmethod
    def last_n_days(n):
        today = DateRange.get_today()
        start_date = today - timedelta(days=n)
        end_date = today
        return start_date, end_date

    @staticmethod
    def this_month():
        today = DateRange.get_today()
        start_date = today.replace(day=1)
        end_date = today
        return start_date, end_date

    @staticmethod
    def this_quarter():
        today = DateRange.get_today()
        current_month = today.month
        start_month = (current_month - 1) // 3 * 3 + 1  # 计算当前季度的开始月份
        start_date = datetime(today.year, start_month, 1).date()

        # 计算季度结束日期
        if start_month == 1:
            end_date = datetime(today.year, 3, 31).date()
        elif start_month == 4:
            end_date = datetime(today.year, 6, 30).date()
        elif start_month == 7:
            end_date = datetime(today.year, 9, 30).date()
        else:  # start_month == 10
            end_date = datetime(today.year, 12, 31).date()

        return start_date, end_date

    @staticmethod
    def this_year():
        today = DateRange.get_today()
        start_date = datetime(today.year, 1, 1).date()
        end_date = today
        return start_date, end_date

    @staticmethod
    def get_range(time_range):

        # 创建时间范围映射字典
        time_range_methods = {
            "this_week": DateRange.this_week,
            "last_week": DateRange.last_week,
            "this_month": DateRange.this_month,
            "this_quarter": DateRange.this_quarter,
            "this_year": DateRange.this_year,
        }

        # 计算日期范围
        if time_range in time_range_methods:
            start_date, end_date = time_range_methods[time_range]()  # 调用相应的方法
        else:
            start_date, end_date = DateRange.this_year()  # 默认返回本年的数据
        return start_date, end_date


# 将 HTML 转换为纯文本的函数，改进 convert_list
def html_to_text(html_content):
    if not html_content:
        return ""
    soup = BeautifulSoup(html_content, "html.parser")

    result = []

    # 处理有序或无序列表，使用递归处理层级
    def convert_list(list_tag, level_prefix="", ordered=True):
        for idx, li in enumerate(list_tag.find_all("li", recursive=False), 1):
            # 确定当前前缀：有序列表用数字前缀，无序列表用符号前缀
            current_prefix = f"{level_prefix}{idx}." if ordered else f"{level_prefix}-"

            # 提取当前 <li> 的直接文本（不包含嵌套的子元素）
            main_text = li.find(text=True, recursive=False)
            main_text = main_text.strip() if main_text else ""
            result.append(f"{current_prefix} {main_text}")

            # 递归处理嵌套的有序或无序列表
            nested_ul = li.find("ul")
            nested_ol = li.find("ol")
            if nested_ul:
                convert_list(nested_ul, level_prefix + "    ", ordered=False)  # 传入当前缩进和无序标记
            elif nested_ol:
                convert_list(nested_ol, level_prefix + "    ", ordered=True)  # 传入当前缩进和有序标记

    # 处理段落和其他标签
    for element in soup.contents:
        if element.name == "ol":
            convert_list(element, ordered=True)  # 有序列表
        elif element.name == "ul":
            convert_list(element, ordered=False)  # 无序列表
        elif element.name == "p" and element.get_text(strip=True):
            result.append(element.get_text(strip=True))
        elif element.name == "strong":
            result.append(f"**{element.get_text(strip=True)}**")  # 加粗的文本
        else:
            result.append(element.get_text(strip=True))

    return "\n".join(result)


# 辅助函数：获取周的起始日期和结束日期
def get_week_date_range(year, week):
    # 根据ISO日历计算周的开始日期（周一）和结束日期（周日）
    # 使用 %G-%V-%u 格式：ISO年份-ISO周数-ISO weekday(1=周一)
    start_date = datetime.strptime(f"{year}-{week:02d}-1", "%G-%V-%u")
    end_date = start_date + timedelta(days=6)
    return f"{start_date.strftime('%Y-%m-%d')}\n-\n{end_date.strftime('%Y-%m-%d')}"


class RecordDownloader:
    @staticmethod
    def download(user_weekly_data, all_weeks, filename):
        """Download weekly report as Excel file.

        Args:
            user_weekly_data: Dict of username -> week -> content
            all_weeks: Set of (year, week) tuples
            filename: Output filename

        Returns:
            Flask send_file response with XLSX attachment
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "软件开发组周报"

        styles = RecordDownloader._setup_workbook_styles()
        RecordDownloader._fill_data(ws, user_weekly_data, all_weeks, styles)
        RecordDownloader._apply_formatting(ws, styles)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    @staticmethod
    def _setup_workbook_styles() -> dict:
        """Create and return style objects for workbook.

        Returns:
            Dict with header_fill, header_font, light_fill, thin_border
        """
        return {
            "header_fill": PatternFill(start_color="808080", end_color="808080", fill_type="solid"),
            "header_font": Font(bold=True, color="FFFFFF"),
            "light_fill": PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid"),
            "thin_border": Border(
                left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
            ),
        }

    @staticmethod
    def _fill_data(ws, user_weekly_data, all_weeks, styles: dict) -> None:
        """Fill headers and data rows in worksheet.

        Args:
            ws: Worksheet to fill
            user_weekly_data: Dict of username -> week -> content
            all_weeks: Set of (year, week) tuples
            styles: Style dict from _setup_workbook_styles
        """
        # Sort weeks and create headers
        sorted_weeks = sorted(all_weeks, key=lambda x: (x[0], x[1]), reverse=True)
        headers = ["姓名"] + [get_week_date_range(year, week) for year, week in sorted_weeks]
        ws.append(headers)

        # Fill data rows
        for username, weekly_data in user_weekly_data.items():
            row = [username]
            for week in sorted_weeks:
                content = weekly_data.get(week, "")
                row.append(html_to_text(content))
            ws.append(row)

    @staticmethod
    def _apply_formatting(ws, styles: dict) -> None:
        """Apply all styling to worksheet.

        Args:
            ws: Worksheet to style
            styles: Style dict from _setup_workbook_styles
        """
        header_fill = styles["header_fill"]
        header_font = styles["header_font"]
        light_fill = styles["light_fill"]
        thin_border = styles["thin_border"]

        # Style header row
        for cell in ws["1"]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Style data rows with alternating colors
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row), start=2):
            for cell in row:
                if row_idx % 2 == 0:
                    cell.fill = light_fill
                cell.alignment = Alignment(horizontal="left", wrap_text=True)

        # Style first column (names)
        for cell in ws["A"]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Set column widths
        for col in ws.columns:
            column = col[0].column_letter
            ws.column_dimensions[column].width = 15

        # Apply borders to all cells
        for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
