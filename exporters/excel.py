"""Excel exporter with rich text cell support.

This module provides the ExcelExporter class for generating XLSX documents
from weekly report records with rich text formatting preserved in cells.
"""

from datetime import datetime, timedelta
from io import BytesIO
from typing import Any

from bs4 import BeautifulSoup, NavigableString
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .base import ExporterBase


class ExcelExporter(ExporterBase):
    """Excel exporter with rich text cell support.

    Generates XLSX documents from weekly report records with:
    - Rich text formatting (bold, italic, underline, strikethrough) preserved
    - Users as rows, weeks as columns structure
    - Professional styling (headers, alternating rows, borders)

    Attributes:
        _uploads_path: Path to uploads directory (for future image support)
    """

    def __init__(self, uploads_path: str | None = None):
        """Initialize with optional uploads path for dependency injection.

        Args:
            uploads_path: Path to uploads directory. If None, will be
                         initialized from Flask current_app.config when needed.
        """
        self._uploads_path = uploads_path

    @property
    def file_extension(self) -> str:
        """Return file extension without dot."""
        return 'xlsx'

    @property
    def mime_type(self) -> str:
        """Return MIME type for send_file()."""
        return 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    def _generate(self, records: list[Any], options: dict) -> BytesIO:
        """Generate XLSX from records using openpyxl.

        Args:
            records: List of Record objects with content and date
            options: Export options (title, etc.)

        Returns:
            BytesIO buffer containing XLSX
        """
        # Group records by user and week (same logic as RecordDownloader)
        user_weekly_data: dict[str, dict[tuple, str]] = {}
        all_weeks: set = set()

        for record in records:
            if not record.content:
                continue

            # Get ISO week info - use strftime for compatibility with mock objects
            date_str = record.date.strftime('%Y-%m-%d')
            record_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            iso_calendar = record_date.isocalendar()
            year, week = iso_calendar[0], iso_calendar[1]
            week_key = (year, week)
            all_weeks.add(week_key)

            # Get user names for this record
            users_found = False
            if record.user:
                for u in record.user:
                    username = u.username
                    if username not in user_weekly_data:
                        user_weekly_data[username] = {}
                    # Combine content if multiple records for same user/week
                    if week_key in user_weekly_data[username]:
                        user_weekly_data[username][week_key] += '<br>' + record.content
                    else:
                        user_weekly_data[username][week_key] = record.content
                    users_found = True

            # Handle case where no users found (e.g., mock objects or missing user)
            if not users_found:
                username = '\u672a\u77e5\u7528\u6237'  # '未知用户'
                if username not in user_weekly_data:
                    user_weekly_data[username] = {}
                if week_key in user_weekly_data[username]:
                    user_weekly_data[username][week_key] += '<br>' + record.content
                else:
                    user_weekly_data[username][week_key] = record.content

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '\u8f6f\u4ef6\u5f00\u53d1\u7ec4\u5468\u62a5'  # '软件开发组周报'

        # Define styles
        header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        light_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Sort weeks by year and week (descending)
        all_weeks = sorted(all_weeks, key=lambda x: (x[0], x[1]), reverse=True)

        # Create headers: ['姓名'] + [week date ranges]
        headers = ['\u59d3\u540d'] + [self._get_week_date_range(year, week) for year, week in all_weeks]
        ws.append(headers)

        # Style header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # Fill data rows
        for username, weekly_data in user_weekly_data.items():
            row = [username]

            for week in all_weeks:
                content = weekly_data.get(week, "")
                if content:
                    # Convert HTML to CellRichText
                    row.append(self._html_to_rich_text(content))
                else:
                    row.append("")

            ws.append(row)

        # Style data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row), start=2):
            for cell in row:
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = light_fill
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Style first column (names) - bold and centered
        for cell in ws['A']:
            if cell.row > 1:  # Skip header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Set column widths
        ws.column_dimensions['A'].width = 12  # Name column
        for col_idx in range(2, ws.max_column + 1):
            col_letter = chr(64 + col_idx)  # B, C, D, etc.
            ws.column_dimensions[col_letter].width = 20

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _get_week_date_range(self, year: int, week: int) -> str:
        """Get week date range string for header.

        Args:
            year: ISO year
            week: ISO week number

        Returns:
            Date range string like "2026-03-23\n-\n2026-03-29"
        """
        # Use ISO calendar format to get start of week (Monday)
        start_date = datetime.strptime(f"{year}-{week:02d}-1", "%G-%V-%u")
        end_date = start_date + timedelta(days=6)
        return f"{start_date.strftime('%Y-%m-%d')}\n-\n{end_date.strftime('%Y-%m-%d')}"

    def _html_to_rich_text(self, html: str):
        """Convert HTML to CellRichText for openpyxl cell value.

        Supports: <strong>, <b>, <em>, <i>, <u>, <s>, <del>, <strike>

        Args:
            html: HTML content string

        Returns:
            CellRichText object with formatting preserved, or plain string
            if no formatting found
        """
        if not html:
            return ""

        soup = BeautifulSoup(html, 'html.parser')
        runs = []

        def process_node(node, styles=None):
            """Recursively process HTML nodes, tracking style state."""
            if styles is None:
                styles = {
                    'bold': False,
                    'italic': False,
                    'underline': None,
                    'strike': False
                }

            # Handle text nodes
            if isinstance(node, NavigableString):
                text = str(node)
                if text.strip():
                    font = InlineFont(
                        b=styles['bold'],
                        i=styles['italic'],
                        u=styles['underline'],  # Must be string 'single' or None
                        strike=styles['strike']
                    )
                    runs.append(TextBlock(font, text))
                return

            # Handle element nodes
            if node.name is None:
                # Root node or document fragment
                for child in node.children:
                    process_node(child, styles)
                return

            # Copy styles for children
            new_styles = styles.copy()

            # Map HTML tags to font properties
            if node.name in ('strong', 'b'):
                new_styles['bold'] = True
            elif node.name in ('em', 'i'):
                new_styles['italic'] = True
            elif node.name == 'u':
                # CRITICAL: underline must be string 'single', not boolean True!
                new_styles['underline'] = 'single'
            elif node.name in ('s', 'strike', 'del'):
                new_styles['strike'] = True
            elif node.name == 'br':
                # Handle line breaks
                runs.append(TextBlock(InlineFont(), '\n'))
                return

            # Process children with updated styles
            for child in node.children:
                process_node(child, new_styles)

        # Process all children of the soup
        for child in soup.children:
            process_node(child)

        # Return CellRichText if we have styled runs, otherwise plain text
        if not runs:
            return soup.get_text()

        return CellRichText(*runs)
